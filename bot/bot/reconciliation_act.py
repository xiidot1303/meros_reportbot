from datetime import date, timedelta

from openpyxl import load_workbook

from bot.bot import *
from app.utils import *
from app.services.client_service import *
from app.services.smartup_service import *
from bot.models import Bot_user, Cabinet


def trim_reconciliation_sheet(file_path: str):
    workbook = load_workbook(file_path)
    if workbook.worksheets:
        worksheet = workbook.active
        if worksheet.max_column > 7:
            worksheet.delete_cols(8, worksheet.max_column - 7)
        workbook.save(file_path)
    return file_path


def get_reconciliation_period(client: Client):
    deferment_days = client.deferment_days
    if deferment_days is None:
        deferment_days = client.secondary_deferment_days
    if deferment_days is None:
        deferment_days = 0

    end_date = date.today()
    start_date = end_date - timedelta(days=deferment_days-1)
    return start_date, end_date


def format_reconciliation_period(context: CustomContext, start_date: date, end_date: date) -> str:
    return context.words.reconciliation_act_period.format(
        start_date=start_date.strftime('%d.%m.%Y'),
        end_date=end_date.strftime('%d.%m.%Y'),
    )


async def _send_reconciliation_act_in_background(context: CustomContext, chat_id: int, client_external_id: str, start_date: date, end_date: date):
    await context.bot.send_chat_action(chat_id=chat_id, action=ChatAction.UPLOAD_DOCUMENT)

    smartup_client = SmartUpApiClient(ApiMethods.reconciliation_act_report)
    reconciliation_act_file_path = smartup_client.reconciliation_act_report(
        client_id=client_external_id,
        start_date=start_date,
        end_date=end_date,
    )
    trim_reconciliation_sheet(reconciliation_act_file_path)

    caption = format_reconciliation_period(context, start_date, end_date)
    with open(reconciliation_act_file_path, 'rb') as file:
        await context.bot.send_document(
            chat_id=chat_id,
            document=file,
            caption=caption,
            reply_markup=await main_menu_keyboard(context),
        )


async def send_reconciliation_act(update: Update, context: CustomContext):
    if update.callback_query:
        await update.callback_query.edit_message_reply_markup(None)

    bot_user: Bot_user = await get_object_by_update(update)
    cabinet: Cabinet = await bot_user.get_active_cabinet
    client: Client = await cabinet.get_client

    start_date, end_date = get_reconciliation_period(client)

    context.application.create_task(
        _send_reconciliation_act_in_background(
            context=context,
            chat_id=update.effective_chat.id,
            client_external_id=client.external_id,
            start_date=start_date,
            end_date=end_date,
        )
    )

    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=context.words.fetching_reconciliation_act
    )
    # send chat action
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action=ChatAction.TYPING)
    return ConversationHandler.END


async def get_start_date(update: Update, context: CustomContext):
    return await send_reconciliation_act(update, context)


async def get_end_date(update: Update, context: CustomContext):
    return await send_reconciliation_act(update, context)
